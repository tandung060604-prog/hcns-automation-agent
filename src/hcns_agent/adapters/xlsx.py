"""Native XLSX parser preserving workbook, cell, formula, and merge structure."""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any

from hcns_agent.domain.canonical import (
    CanonicalDocument,
    DocumentContent,
    ParserProvenance,
    Sheet,
    SourceLocation,
    SpreadsheetCell,
    SpreadsheetRow,
    Workbook,
)
from hcns_agent.domain.documents import ParseStatus, SourceFormat
from hcns_agent.domain.errors import DocumentIntakeError, IntakeErrorCode
from hcns_agent.ports.document_parser import (
    DocumentSource,
    ParseContext,
    ParserCapabilities,
)


class XlsxDocumentParser:
    name = "xlsx/openpyxl-native"
    version = "1.0.0"
    capabilities = ParserCapabilities(
        source_formats=frozenset({SourceFormat.XLSX}),
        preserves_tables=True,
    )

    def __init__(self, *, maximum_cells: int = 1_000_000) -> None:
        self._maximum_cells = maximum_cells

    def supports(self, source_format: SourceFormat) -> bool:
        return source_format in self.capabilities.source_formats

    def parse(self, source: DocumentSource, context: ParseContext) -> CanonicalDocument:
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
            from openpyxl.cell.cell import MergedCell  # type: ignore[import-untyped]

            native_workbook = load_workbook(
                filename=BytesIO(source.content),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as error:
            raise DocumentIntakeError(
                IntakeErrorCode.CORRUPTED_FILE,
                "XLSX workbook could not be parsed",
            ) from error

        sheets: list[Sheet] = []
        try:
            for sheet_index, worksheet in enumerate(native_workbook.worksheets):
                if worksheet.max_row * worksheet.max_column > self._maximum_cells:
                    raise DocumentIntakeError(
                        IntakeErrorCode.FILE_TOO_LARGE,
                        "XLSX worksheet exceeds the configured cell limit",
                    )
                rows: list[SpreadsheetRow] = []
                for row_index, native_row in enumerate(worksheet.iter_rows(), start=1):
                    cells: list[SpreadsheetCell] = []
                    for native_cell in native_row:
                        if isinstance(native_cell, MergedCell) or native_cell.value is None:
                            continue
                        raw_value = native_cell.value
                        formula = (
                            str(raw_value)
                            if native_cell.data_type == "f"
                            or (isinstance(raw_value, str) and raw_value.startswith("="))
                            else None
                        )
                        cells.append(
                            SpreadsheetCell(
                                row_index=native_cell.row,
                                column_index=native_cell.column,
                                coordinate=native_cell.coordinate,
                                value=_scalar(raw_value),
                                data_type=str(native_cell.data_type),
                                formula=formula,
                                source=SourceLocation(
                                    source_reference=source.source_reference,
                                    sheet_name=worksheet.title,
                                    row_index=native_cell.row,
                                    column_index=native_cell.column,
                                ),
                            )
                        )
                    if cells:
                        rows.append(SpreadsheetRow(row_index=row_index, cells=tuple(cells)))
                sheets.append(
                    Sheet(
                        name=worksheet.title,
                        index=sheet_index,
                        rows=tuple(rows),
                        merged_ranges=tuple(
                            str(cell_range) for cell_range in worksheet.merged_cells.ranges
                        ),
                    )
                )
        finally:
            native_workbook.close()

        return CanonicalDocument(
            document_id=source.document_id,
            source=context.source_descriptor,
            source_format=context.source_format,
            content=DocumentContent(workbook=Workbook(sheets=tuple(sheets))),
            parse_status=ParseStatus.SUCCESS,
            provenance=(
                ParserProvenance(
                    parser_name=self.name,
                    parser_version=self.version,
                    source_format=context.source_format,
                    metadata={"library": "openpyxl", "formulaMode": "preserve"},
                ),
            ),
        )


def _scalar(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)
